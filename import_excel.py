import openpyxl
import json
import re
import urllib.request
import sys
import os

F = r'C:\Users\do\Desktop\毛线库存表\毛线库存管理表(1).xlsx'
API = 'http://localhost:3001/api'  # 改为实际服务器地址

CATEGORY_RULES = {
    '蛋糕线': {'keywords': ['蛋糕线'], 'defaultWeight': 0, 'defaultUnit': '袋'},
    '冰条线(一团)': {'keywords': ['一团冰条', '冰条(一团', '1团冰条', '01(一团冰条'], 'defaultWeight': 0.25, 'defaultUnit': '团'},
    '冰条线(两团)': {'keywords': ['两团冰条', '冰条(两团', '2团冰条', '两团 冰条', '2团 冰条'], 'defaultWeight': 0.5, 'defaultUnit': '团'},
    '冰条线(三团)': {'keywords': ['三团冰条', '冰条(三团', '3团冰条', '四团冰条', '冰条(四团', '4团冰条'], 'defaultWeight': 0.75, 'defaultUnit': '团'},
    '羊毛线': {'keywords': ['羊毛线'], 'defaultWeight': 0.05, 'defaultUnit': '个'},
    '抹布': {'keywords': ['抹布'], 'defaultWeight': 0.25, 'defaultUnit': '卷'},
    '普通毛线': {'keywords': [], 'defaultWeight': 14.4, 'defaultUnit': '箱'},
}


def clean_model(model):
    if not model:
        return ''
    model = str(model).strip()
    model = model.replace('＃', '#').replace('（', '(').replace('）', ')')
    model = model.replace('#', '')
    model = re.sub(r'\s+', '', model)
    return model


def clean_color(color):
    if not color:
        return ''
    return str(color).strip()


def clean_date(date_str):
    if not date_str:
        return ''
    s = str(date_str).strip()
    if s.count('.') == 2:
        parts = s.split('.')
        return f'{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}'
    if '-' in s and len(s.split('-')) == 3:
        return s
    m = re.match(r'(\d{4})[/.](\d{1,2})[/.](\d{1,2})', s)
    if m:
        return f'{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}'
    return s


def parse_qty(val):
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r'[包袋团个卷箱]', '', s)
    try:
        return float(s)
    except ValueError:
        return None


def classify_product(model, color):
    model_str = clean_model(model) + ' ' + (clean_color(color) or '')
    for cat, rule in CATEGORY_RULES.items():
        if cat == '普通毛线':
            continue
        for kw in rule['keywords']:
            if kw in model_str:
                return cat
    if re.match(r'^\d{2,3}$', model_str.strip().split()[0] if model_str.strip() else ''):
        return '普通毛线'
    return '其他'


def guess_product_info(model, color, sheet_kg):
    cat = classify_product(model, color)
    rule = CATEGORY_RULES.get(cat, CATEGORY_RULES['其他'])
    weight = sheet_kg if sheet_kg else rule['defaultWeight']
    unit = rule['defaultUnit']

    if cat in ('普通毛线',):
        weight = 14.4
        if sheet_kg and sheet_kg != 14.4:
            weight = sheet_kg

    return cat, weight, unit


def process_inbound_sheet(ws):
    print(f'处理入库明细，共 {ws.max_row} 行...')
    products = {}
    transactions = []
    problems = []
    
    for row_idx in range(5, ws.max_row + 1):
        seq = ws.cell(row=row_idx, column=1).value
        if not seq:
            continue
        date = clean_date(ws.cell(row=row_idx, column=2).value)
        model = clean_model(ws.cell(row=row_idx, column=3).value)
        color = clean_color(ws.cell(row=row_idx, column=4).value)
        kg_per_unit = ws.cell(row=row_idx, column=5).value
        qty_raw = ws.cell(row=row_idx, column=6).value
        price = ws.cell(row=row_idx, column=8).value
        operator = ws.cell(row=row_idx, column=10).value

        if not model:
            if color:
                problems.append(f'入库行{row_idx}: 颜色={color}, 型号为空, 跳过')
            continue

        qty = parse_qty(qty_raw)
        if qty is None:
            problems.append(f'入库行{row_idx}: {model} {color}, 无法解析数量: {qty_raw}')
            continue

        kg_val = None
        try:
            kg_val = float(ws.cell(row=row_idx, column=7).value or 0)
        except:
            pass

        cat, wt, unit = guess_product_info(model, color, kg_per_unit)
        if not wt:
            wt = 0

        key = f'{model}|{color}'
        if key not in products:
            units = []
            if cat == '冰条线(一团)':
                units = [{'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(两团)':
                units = [{'unit': '包', 'weight': 0.5}, {'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(三团)':
                units = [{'unit': '包', 'weight': 0.75}, {'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(四团)':
                units = [{'unit': '包', 'weight': 1.0}, {'unit': '团', 'weight': 0.25}]
            elif cat == '普通毛线':
                units = [{'unit': '箱', 'weight': 14.4}, {'unit': '袋', 'weight': 1.8}]
            elif cat == '羊毛线':
                units = [{'unit': '个', 'weight': 0.05}]
            elif cat == '抹布':
                units = [{'unit': '卷', 'weight': 0.25}]
            else:
                units = [{'unit': unit, 'weight': wt}]
            
            products[key] = {
                '_key': key, 'cat': cat, 'model': model, 'color': color,
                'units': units, 'safety': 5
            }

        calc_kg = qty * wt if wt else (kg_val or 0)
        if not calc_kg and kg_val:
            calc_kg = kg_val
        if not calc_kg:
            calc_kg = 0

        price_val = 0
        try:
            price_val = float(price or 0)
        except:
            pass

        txn = {
            '_importKey': key, 'type': 'in', 'date': date,
            'qty': qty, 'unit': products[key]['units'][0]['unit'],
            'unitwt': wt, 'kg': calc_kg,
            'price': price_val, 'amt': calc_kg * price_val,
            'customer': '', 'operator': str(operator or '').strip(),
            'notes': '', 'outType': ''
        }
        transactions.append(txn)

    print(f'  产品: {len(products)}, 记录: {len(transactions)}, 问题: {len(problems)}')
    return products, transactions, problems


def process_outbound_sheet(ws):
    print(f'处理出库明细，共 {ws.max_row} 行...')
    products = {}
    transactions = []
    problems = []

    for row_idx in range(5, ws.max_row + 1):
        seq = ws.cell(row=row_idx, column=1).value
        if not seq:
            continue
        date = clean_date(ws.cell(row=row_idx, column=2).value)
        model = clean_model(ws.cell(row=row_idx, column=3).value)
        color = clean_color(ws.cell(row=row_idx, column=4).value)
        kg_per_unit = ws.cell(row=row_idx, column=5).value
        qty_raw = ws.cell(row=row_idx, column=6).value
        customer = ws.cell(row=row_idx, column=10).value
        address = ws.cell(row=row_idx, column=11).value

        if not model:
            if color:
                problems.append(f'出库行{row_idx}: 颜色={color}, 型号为空, 跳过')
            continue

        qty = parse_qty(qty_raw)
        if qty is None:
            problems.append(f'出库行{row_idx}: {model} {color}, 无法解析数量: {qty_raw}')
            continue

        cat, wt, unit = guess_product_info(model, color, kg_per_unit)
        if not wt:
            wt = 0

        key = f'{model}|{color}'
        if key not in products:
            if cat == '冰条线(一团)':
                units = [{'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(两团)':
                units = [{'unit': '包', 'weight': 0.5}, {'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(三团)':
                units = [{'unit': '包', 'weight': 0.75}, {'unit': '团', 'weight': 0.25}]
            elif cat == '冰条线(四团)':
                units = [{'unit': '包', 'weight': 1.0}, {'unit': '团', 'weight': 0.25}]
            elif cat == '普通毛线':
                units = [{'unit': '箱', 'weight': 14.4}, {'unit': '袋', 'weight': 1.8}]
            elif cat == '羊毛线':
                units = [{'unit': '个', 'weight': 0.05}]
            else:
                units = [{'unit': unit, 'weight': wt}]
            
            products[key] = {
                '_key': key, 'cat': cat, 'model': model, 'color': color,
                'units': units, 'safety': 5
            }

        calc_kg = qty * wt
        if not calc_kg:
            calc_kg = 0

        cust = str(customer or '').strip()
        notes = str(address or '').strip()

        txn = {
            '_importKey': key, 'type': 'out', 'date': date,
            'qty': qty, 'unit': products[key]['units'][0]['unit'],
            'unitwt': wt, 'kg': calc_kg,
            'price': 0, 'amt': 0,
            'customer': cust, 'operator': '',
            'notes': notes, 'outType': '销售出库'
        }
        transactions.append(txn)

    print(f'  产品: {len(products)}, 记录: {len(transactions)}, 问题: {len(problems)}')
    return products, transactions, problems


def merge_products(p1, p2):
    merged = {}
    for k, v in p1.items():
        merged[k] = v
    for k, v in p2.items():
        if k in merged:
            if v['units'] and not merged[k]['units']:
                merged[k]['units'] = v['units']
            if v['color'] and not merged[k]['color']:
                merged[k]['color'] = v['color']
        else:
            merged[k] = v
    return list(merged.values())


def main():
    print('读取Excel...')
    wb = openpyxl.load_workbook(F)
    
    in_ws = wb['入库明细']
    out_ws = None
    for name in wb.sheetnames:
        if '出库' in name:
            out_ws = wb[name]
            break

    in_prods, in_txns, in_probs = process_inbound_sheet(in_ws)

    if out_ws:
        out_prods, out_txns, out_probs = process_outbound_sheet(out_ws)
    else:
        out_prods, out_txns, out_probs = {}, [], []

    all_prods = merge_products(in_prods, out_prods)
    all_txns = in_txns + out_txns
    all_probs = in_probs + out_probs

    customers = list(set(t['customer'] for t in all_txns if t['customer']))
    categories = ['普通毛线', '蛋糕线', '冰条线(一团)', '冰条线(两团)', '冰条线(三团)', '冰条线(四团)', '羊毛线', '抹布', '其他']

    payload = {
        'prods': all_prods,
        'txns': all_txns,
        'categories': categories,
        'customers': customers
    }

    print(f'\n=== 导入汇总 ===')
    print(f'产品数: {len(all_prods)}')
    print(f'记录数: {len(all_txns)}')
    print(f'入库记录: {len(in_txns)}, 出库记录: {len(out_txns)}')
    print(f'客户数: {len(customers)}')
    print(f'问题行: {len(all_probs)}')

    if all_probs:
        prob_file = os.path.join(os.path.dirname(F), '导入问题清单.txt')
        with open(prob_file, 'w', encoding='utf-8') as f:
            f.write('Excel导入问题清单\n')
            f.write('=' * 50 + '\n')
            for p in all_probs:
                f.write(p + '\n')
        print(f'问题清单已保存: {prob_file}')

    json_file = os.path.join(os.path.dirname(F), 'import_data.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f'导入数据已保存: {json_file}')

    print(f'\n请确认后通过API导入: curl -X POST {API}/import -H "Content-Type: application/json" -d @{json_file}')


if __name__ == '__main__':
    main()
